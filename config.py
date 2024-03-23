import os

import torch
import whisper
from dotenv import load_dotenv, find_dotenv
import openpyxl
from openai import OpenAI

load_dotenv(dotenv_path=find_dotenv('.env'))

FILE_NAME = str(os.getenv('EXCEL_FILE'))
workbook = openpyxl.load_workbook(FILE_NAME)
URL: str = os.getenv('API_URL')
API_KEY: str = os.getenv('API_KEY')
CHAT_URL: str = os.getenv('CHAT_URL')

client = OpenAI(api_key=API_KEY)

AUDIOS_FOLDER = "audios/"
AUDIOS_QUESTIONNAIRE_FOLDER = "../опитування"

EMOTIONS_API_KEY: str = os.getenv('EMOTIONS_API_KEY')
EMOTIONS_API_KEY_PASSWORD: str = os.getenv('EMOTIONS_API_KEY_PASSWORD')
SPLIT_AUDIOS = "splits/"
RESULT_EMOTIONS_FILE = "emotions_results.json"
API_EMOTIONS_URL = "https://cloud.emlo.cloud/analysis/analyzeFile"

torch.cuda.is_available()
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
model = whisper.load_model("medium", device=DEVICE)

FILE_QUESTIONNAIRE = '../questionnaire.xlsx'
workbook_questionnaire = openpyxl.load_workbook(FILE_QUESTIONNAIRE)

FILE_EMOTIONS = '../emotions.xlsx'
workbook_emotions = openpyxl.load_workbook(FILE_EMOTIONS)
